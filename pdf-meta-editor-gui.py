# -*- coding: utf-8 -*-
import os
import sys
import re  # Import regular expressions
import openpyxl  # Now needed again
from PySide6.QtWidgets import (
    QApplication,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QFileDialog,
    QMessageBox,
    QProgressBar,
    QComboBox,
    QSpacerItem,
    QSizePolicy,
    QScrollArea,
    QCheckBox,
    QGroupBox,
    QTableView,
    QHeaderView,
    QSpinBox,
)
from PySide6.QtCore import Qt, QThread, Signal, QAbstractTableModel, QModelIndex, QEvent
from PySide6.QtGui import QStandardItemModel, QStandardItem, QKeySequence

from pypdf import PdfReader, PdfWriter
from pypdf.errors import PdfReadError
from pypdf.generic import NameObject, TextStringObject


# --- PDF Metadata Processing Logic (set_pdf_metadata) ---
# (This function remains unchanged from the previous version)
def set_pdf_metadata(
    input_pdf_path: str,
    output_pdf_path: str,
    new_title: str | None = None,
    new_author: str | None = None,
    new_subject: str | None = None,
    new_creator: str | None = None,
    new_producer: str | None = None,
    new_keywords: str | None = None,
    set_initial_view_defaults: bool = True,
):
    """
    Edits specified metadata fields and sets initial view defaults of a PDF file.
    (Docstring omitted for brevity, same as original)
    """
    try:
        reader = PdfReader(input_pdf_path, strict=False)
    except FileNotFoundError:
        raise  # Re-raise FileNotFoundError to be caught by worker
    except PdfReadError as e:
        print(
            f"Warning: Corrupt or unreadable PDF: {os.path.basename(input_pdf_path)} - {e}"
        )
        raise  # Re-raise PdfReadError
    except Exception as e:
        print(
            f"Warning: Error reading PDF: {os.path.basename(input_pdf_path)} - {e}"
        )
        raise  # Re-raise other read errors

    writer = PdfWriter()
    try:
        writer.clone_document_from_reader(reader)
    except Exception as e:
        # Handle potential issues during cloning, e.g., complex structures
        print(
            f"Warning: Could not fully clone document: {os.path.basename(input_pdf_path)} - {e}"
        )
        # Attempt to add pages individually as a fallback
        try:
            writer.append_pages_from_reader(reader)
        except Exception as e_append:
            print(
                f"Error: Failed to append pages for {os.path.basename(input_pdf_path)}: {e_append}"
            )
            raise Exception(
                f"Cloning and appending failed for {input_pdf_path}"
            ) from e_append

    metadata_update = {}
    # Map internal names to PDF dictionary keys
    field_map = {
        "/Title": new_title,
        "/Author": new_author,
        "/Subject": new_subject,
        "/Creator": new_creator,
        "/Producer": new_producer,
        "/Keywords": new_keywords,
    }

    for key, value in field_map.items():
        if value is not None:  # Process if not "Leave Unchanged"
            metadata_update[key] = TextStringObject(value)  # Handles "" for clear

    # Get existing metadata
    merged_metadata = {}
    try:
        existing_metadata = reader.metadata
        if existing_metadata:
            merged_metadata.update(existing_metadata)
    except Exception:
        print(
            f"Warning: Could not read existing metadata for {os.path.basename(input_pdf_path)}"
        )

    # Apply updates (overwrites existing or sets empty)
    merged_metadata.update(metadata_update)

    if merged_metadata:
        try:
            writer.add_metadata(merged_metadata)
        except Exception as e:
            print(
                f"Warning: Could not write metadata for {os.path.basename(input_pdf_path)}: {e}"
            )

    # Set initial view defaults
    if set_initial_view_defaults:
        try:
            if writer.root_object and isinstance(writer.root_object, dict):
                # 1. Set Navigation Tab to "Page Only"
                writer.root_object[NameObject("/PageMode")] = NameObject("/UseNone")

                # 2. Remove Page Layout setting (use default)
                if NameObject("/PageLayout") in writer.root_object:
                    del writer.root_object[NameObject("/PageLayout")]

                # 3. Remove OpenAction (often controls initial zoom/destination)
                if NameObject("/OpenAction") in writer.root_object:
                    del writer.root_object[NameObject("/OpenAction")]

            else:
                print(
                    f"Warning: Could not set initial view defaults for {os.path.basename(input_pdf_path)} - Root object issue."
                )
        except Exception as e:
            print(
                f"Warning: Could not set initial view defaults for {os.path.basename(input_pdf_path)}: {e}"
            )

    # Write output
    os.makedirs(os.path.dirname(output_pdf_path), exist_ok=True)
    try:
        with open(output_pdf_path, "wb") as output_file:
            writer.write(output_file)
    except Exception as e:
        print(
            f"Error: Could not write output file {os.path.basename(output_pdf_path)}: {e}"
        )
        # Attempt to remove potentially corrupted partial file
        if os.path.exists(output_pdf_path):
            try:
                os.remove(output_pdf_path)
            except OSError:
                pass  # Ignore removal error
        raise  # Re-raise write error

    return True


# --- Combined Worker Thread (Uses Masked Base Name) ---
# (This class remains unchanged from the previous version - no Excel logic needed here)
class CombinedWorker(QThread):
    progress = Signal(int)
    log_message = Signal(str)
    # Signal now reports: planned_renames, edited_succeeded, edited_failed
    finished = Signal(int, int, int)

    def __init__(
        self,
        pdf_folder,
        output_folder,
        # Data now includes original name, masked base, and append text
        rename_data,  # List of tuples: [(original_filename, masked_base_name, text_to_append), ...]
        metadata_settings,
        parent=None,
    ):
        super().__init__(parent)
        self.pdf_folder = pdf_folder
        self.output_folder = output_folder
        self.rename_data = rename_data
        self.metadata_settings = metadata_settings
        self.is_running = True

    def run(self):
        planned_renames = 0
        edited_succeeded = 0
        edited_failed = 0
        total_steps = 0
        current_step = 0

        self.log_message.emit("Starting process...")

        # Get count from rename_data
        total_files = len(self.rename_data)

        if total_files == 0:
            self.log_message.emit("No PDF files loaded for processing.")
            self.finished.emit(0, 0, 0)
            return

        total_steps = total_files * 2  # 1 for name planning, 1 for processing

        # --- Phase 1: Determine Target Filenames (Using Masked Base + Append Text) ---
        self.log_message.emit("--- Determining Target Filenames ---")
        target_filenames_map = {}  # original_filename -> target_output_filename

        for original_filename, masked_base_name, text_to_append in self.rename_data:
            if not self.is_running:
                break
            current_step += 1

            target_filename = (
                original_filename
            )  # Default: use original if no append text or masked name invalid
            _, original_ext = os.path.splitext(original_filename)  # Get extension from original

            # Use the provided masked_base_name if it's valid (not empty)
            # If masked_base_name is empty, it implies the mask failed, so we use original
            if masked_base_name:
                if text_to_append and text_to_append.strip():
                    text_to_append_clean = text_to_append.strip()
                    # Construct target using the provided masked_base_name
                    potential_new_name = (
                        f"{masked_base_name} {text_to_append_clean}{original_ext}"
                    )
                else:
                    # No append text, but valid mask - use masked name + original extension
                    potential_new_name = f"{masked_base_name}{original_ext}"

                # Check if the final name is actually different from the original
                if potential_new_name != original_filename:
                    target_filename = potential_new_name
                    planned_renames += 1
                    self.log_message.emit(
                        f"  Planning rename: '{original_filename}' -> '{target_filename}'"
                    )
                else:
                    # If result is same name, treat as no rename planned
                    target_filename = original_filename
            else:
                # Masked base name was empty (mask failed), target is the original name
                target_filename = original_filename

            target_filenames_map[original_filename] = target_filename

            if total_steps > 0:
                self.progress.emit(int((current_step / total_steps) * 100))

        self.log_message.emit(
            f"Filename planning complete. {planned_renames} file(s) will have different output names."
        )

        # --- Phase 2: Metadata Editing & Writing with Target Name ---
        if not self.is_running:
            self.log_message.emit("--- Process stopped before Metadata Phase ---")
            self.finished.emit(planned_renames, 0, 0)
            return

        self.log_message.emit("--- Starting Metadata Editing and Writing Phase ---")
        os.makedirs(self.output_folder, exist_ok=True)

        # Iterate using original filename from rename_data to ensure we process all intended files
        for original_filename, _, _ in self.rename_data:  # We only need original name as key here
            if not self.is_running:
                break
            current_step += 1

            target_filename = target_filenames_map.get(
                original_filename, original_filename
            )
            input_path = os.path.join(self.pdf_folder, original_filename)
            # Handle case where input file might have been renamed *if* outputting in place
            # This shouldn't happen with the current logic (we read original, write target)
            # but adding a check for safety if output == input.
            if (
                self.output_folder == self.pdf_folder
                and not os.path.exists(input_path)
            ):
                # Check if the target name exists (maybe it was renamed in a previous failed run?)
                potential_input_path = os.path.join(
                    self.pdf_folder, target_filename
                )
                if os.path.exists(potential_input_path):
                    input_path = potential_input_path
                    self.log_message.emit(
                        f"  Warning: Original '{original_filename}' not found, using existing '{target_filename}' as input."
                    )
                else:
                    self.log_message.emit(
                        f"  -> Error: Input file '{original_filename}' not found (skipped)."
                    )
                    edited_failed += 1
                    if total_steps > 0:
                        self.progress.emit(int((current_step / total_steps) * 100))
                    continue

            output_path = os.path.join(self.output_folder, target_filename)

            self.log_message.emit(
                f"Processing: '{original_filename}' -> Saving as: '{target_filename}'..."
            )

            if not os.path.exists(input_path):
                self.log_message.emit(
                    f"  -> Error: Input file '{original_filename}' not found (skipped)."
                )
                edited_failed += 1
                if total_steps > 0:
                    self.progress.emit(int((current_step / total_steps) * 100))
                continue

            # --- Calculate Metadata ---
            target_filename_base = os.path.splitext(target_filename)[0]
            values_to_set = {}
            for field_key, settings in self.metadata_settings.items():
                mode = settings["mode"]
                specific_value = settings["value"]
                current_value = None
                if mode == "Specific":
                    current_value = specific_value
                elif mode == "Clear":
                    current_value = ""
                elif mode == "Filename" and field_key == "title":
                    current_value = target_filename_base
                elif mode == "FilenameAfterSpace" and field_key == "title":
                    first_space_index = target_filename_base.find(" ")
                    if first_space_index != -1:
                        current_value = target_filename_base[
                            first_space_index + 1 :
                        ].strip()
                    else:
                        current_value = target_filename_base
                    if not current_value:
                        current_value = target_filename_base
                values_to_set[field_key] = current_value

            # --- Apply Metadata and Write Output ---
            try:
                set_pdf_metadata(
                    input_path,
                    output_path,
                    new_title=values_to_set.get("title"),
                    new_author=values_to_set.get("author"),
                    new_subject=values_to_set.get("subject"),
                    new_creator=values_to_set.get("creator"),
                    new_producer=values_to_set.get("producer"),
                    new_keywords=values_to_set.get("keywords"),
                    set_initial_view_defaults=True,
                )
                log_parts = []
                for field, value in values_to_set.items():
                    if value is None:
                        log_status = "(unchanged)"
                    elif value == "":
                        log_status = "(cleared)"
                    else:
                        log_status = f"'{value}'"
                    log_parts.append(f"{field.capitalize()}={log_status}")
                self.log_message.emit(
                    f"  -> Success: Metadata set. Saved to '{target_filename}' in output."
                )
                edited_succeeded += 1
            except FileNotFoundError:
                self.log_message.emit(
                    f"  -> Error: Input file '{original_filename}' disappeared? (skipped)"
                )
                edited_failed += 1
            except PdfReadError as e:
                self.log_message.emit(
                    f"  -> Error: Cannot read PDF '{original_filename}' (skipped): {e}"
                )
                edited_failed += 1
            except Exception as e:
                self.log_message.emit(
                    f"  -> Error: Processing/writing failed for '{original_filename}' -> '{target_filename}': {e}"
                )
                edited_failed += 1

            if total_steps > 0:
                self.progress.emit(int((current_step / total_steps) * 100))

        # --- Phase 3: Completion ---
        status_msg = (
            "Processing finished" if self.is_running else "Processing stopped by user"
        )
        self.log_message.emit("--------------------")
        self.log_message.emit(
            f"{status_msg}."
            f" Renames Planned: {planned_renames}."
            f" Files Processed (Metadata): {edited_succeeded}✓ {edited_failed}✗."
        )
        if self.is_running and total_steps > 0:
            self.progress.emit(100)
        self.finished.emit(planned_renames, edited_succeeded, edited_failed)

    def stop(self):
        self.log_message.emit("Stop signal received...")
        self.is_running = False


# --- Custom TableView with Enhanced Interaction ---
# (This class remains unchanged from the previous version)
class PastingTableView(QTableView):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Enable extended selection (drag, shift-click, ctrl-click)
        self.setSelectionMode(QTableView.SelectionMode.ExtendedSelection)

    def keyPressEvent(self, event):
        model = self.model()
        if not model:
            super().keyPressEvent(event)
            return

        selected_indexes = self.selectionModel().selectedIndexes()

        # --- Handle Paste ---
        if event.matches(QKeySequence.StandardKey.Paste):
            if not selected_indexes:
                event.accept()
                return  # No target for paste

            # Check if the target column (col 1) is editable
            target_col = 1
            first_selected_row = min(idx.row() for idx in selected_indexes)
            test_index = model.index(first_selected_row, target_col)
            if not (model.flags(test_index) & Qt.ItemFlag.ItemIsEditable):
                QMessageBox.information(
                    self,
                    "Paste Disabled",
                    "Pasting into the 'Text to Append' column is disabled when 'Append from Excel' mode is active.",
                )
                event.accept()
                return  # Don't paste if not editable

            clipboard = QApplication.clipboard()
            clipboard_text = clipboard.text()
            if not clipboard_text:
                event.accept()
                return

            # Use the top-left cell of the selection as the starting point
            start_row = min(index.row() for index in selected_indexes)
            # Force pasting into the second column (index 1)
            start_col = 1

            lines = clipboard_text.strip("\n").split("\n")
            model.blockSignals(True)
            try:
                current_row = start_row
                for r, line in enumerate(lines):
                    if current_row >= model.rowCount():
                        break
                    cells = line.split("\t")
                    if cells:
                        paste_value = cells[0].strip()
                        target_index = model.index(current_row, start_col)
                        # Double-check flags just in case
                        flags = model.flags(target_index)
                        if flags & Qt.ItemFlag.ItemIsEditable:
                            model.setData(
                                target_index, paste_value, Qt.EditRole
                            )
                    current_row += 1
            except Exception as e:
                print(f"Error during paste operation: {e}")
            finally:
                model.blockSignals(False)
                # Trigger dataChanged signal manually if needed, though setData should do it
                # model.dataChanged.emit(model.index(start_row, start_col), model.index(current_row - 1, start_col))
            event.accept()

        # --- Handle Delete ---
        elif event.key() == Qt.Key_Delete:
            if not selected_indexes:
                event.accept()
                return

            # Check if the target column (col 1) is editable
            target_col = 1
            first_selected_row = min(idx.row() for idx in selected_indexes)
            test_index = model.index(first_selected_row, target_col)
            if not (model.flags(test_index) & Qt.ItemFlag.ItemIsEditable):
                QMessageBox.information(
                    self,
                    "Delete Disabled",
                    "Deleting from the 'Text to Append' column is disabled when 'Append from Excel' mode is active.",
                )
                event.accept()
                return  # Don't delete if not editable

            model.blockSignals(True)
            try:
                # Important: Iterate over a copy of the list or unique indexes
                # as modifying the model might affect the selection model implicitly
                unique_rows_cols = set(
                    (idx.row(), idx.column()) for idx in selected_indexes
                )
                min_row_deleted = model.rowCount()
                max_row_deleted = -1
                for r, c in unique_rows_cols:
                    # Only delete from the editable column (column 1)
                    if c == 1:
                        index_to_delete = model.index(r, c)
                        flags = model.flags(index_to_delete)
                        if flags & Qt.ItemFlag.ItemIsEditable:
                            model.setData(index_to_delete, "", Qt.EditRole)  # Set to empty string
                            min_row_deleted = min(min_row_deleted, r)
                            max_row_deleted = max(max_row_deleted, r)
            except Exception as e:
                print(f"Error during delete operation: {e}")
            finally:
                model.blockSignals(False)
                # Emit dataChanged for the affected range
                # if min_row_deleted <= max_row_deleted:
                #     model.dataChanged.emit(model.index(min_row_deleted, target_col), model.index(max_row_deleted, target_col))

            event.accept()

        # --- Handle Copy ---
        elif event.matches(QKeySequence.StandardKey.Copy):
            if not selected_indexes:
                event.accept()
                return

            # Determine range of selection (min/max row and col)
            min_row = min(index.row() for index in selected_indexes)
            max_row = max(index.row() for index in selected_indexes)
            # We only copy from column 1
            min_col = 1
            max_col = 1

            copied_lines = []
            for r in range(min_row, max_row + 1):
                index_to_copy = model.index(r, min_col)  # Always column 1
                # Check if this specific cell was actually selected
                # (Needed for non-contiguous Ctrl+Click selections)
                # A simpler approach for drag/shift selection is to just copy the whole block
                is_selected = self.selectionModel().isSelected(index_to_copy)
                if is_selected:
                    cell_data = model.data(index_to_copy, Qt.DisplayRole)
                    copied_lines.append(
                        str(cell_data) if cell_data is not None else ""
                    )
                else:
                    # Handle non-contiguous selection - add empty string or skip?
                    # Adding empty string maintains row structure if user expects it
                    copied_lines.append("")

            # Join lines with newline
            copied_text = "\n".join(copied_lines)

            if copied_text:
                clipboard = QApplication.clipboard()
                clipboard.setText(copied_text)

            event.accept()

        # --- Default Handling ---
        else:
            # Allow default table navigation, editing triggers, etc.
            super().keyPressEvent(event)


# --- Main Application Window (Adds Excel Option) ---
class CombinedPdfProcessorApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PDF Renamer & Metadata Editor")
        self.setGeometry(100, 100, 800, 980)  # Increased width/height slightly
        self.worker = None
        self.excel_mapping = {}  # To store loaded drawing_num -> title mapping

        # --- Main Layout ---
        main_layout = QVBoxLayout(self)

        # --- File/Folder Selection Group ---
        # (This section remains unchanged)
        file_group = QGroupBox("Files and Folders")
        file_layout = QVBoxLayout()
        pdf_folder_layout = QHBoxLayout()
        self.pdf_folder_label = QLabel("PDF Folder (Input):")
        self.pdf_folder_input = QLineEdit()
        self.pdf_folder_button = QPushButton("Select...")
        self.pdf_folder_button.clicked.connect(self.select_pdf_folder)
        pdf_folder_layout.addWidget(self.pdf_folder_label)
        pdf_folder_layout.addWidget(self.pdf_folder_input)
        pdf_folder_layout.addWidget(self.pdf_folder_button)
        file_layout.addLayout(pdf_folder_layout)
        out_folder_layout = QHBoxLayout()
        self.out_folder_label = QLabel("Output Folder (Processed Files):")
        self.out_folder_input = QLineEdit()
        self.out_folder_input.setPlaceholderText("Default: Input Folder + '_processed'")
        self.out_folder_button = QPushButton("Select...")
        self.out_folder_button.clicked.connect(self.select_output_folder)
        self.edit_in_place_checkbox = QCheckBox("Output to Input Folder")
        self.edit_in_place_checkbox.setToolTip(
            "Check this to save processed files directly in the PDF Input Folder (potentially overwriting)."
        )
        self.edit_in_place_checkbox.stateChanged.connect(self.toggle_output_folder)
        out_folder_layout.addWidget(self.out_folder_label)
        out_folder_layout.addWidget(self.out_folder_input)
        out_folder_layout.addWidget(self.out_folder_button)
        file_layout.addLayout(out_folder_layout)
        file_layout.addWidget(
            self.edit_in_place_checkbox, alignment=Qt.AlignmentFlag.AlignRight
        )
        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)

        # --- Renaming Table Group (Adds Mode Selection and Excel Options) ---
        self.rename_group = QGroupBox("Rename Files on Output (Optional)")
        rename_layout = QVBoxLayout()

        # -- Mode Selection --
        mode_select_layout = QHBoxLayout()
        mode_select_layout.addWidget(QLabel("Append Text Source:"))
        self.rename_mode_combo = QComboBox()
        self.rename_mode_combo.addItems(["Manual / Paste Table", "Append from Excel"])
        self.rename_mode_combo.currentIndexChanged.connect(self._update_rename_mode)
        mode_select_layout.addWidget(self.rename_mode_combo)
        mode_select_layout.addStretch()
        rename_layout.addLayout(mode_select_layout)

        # --- Regex Masking Section ---
        regex_layout = QVBoxLayout()

        # --- Regex Selection Dropdown ---
        regex_select_layout = QHBoxLayout()
        regex_select_layout.addWidget(QLabel("Pre-made Regex Pattern:"))
        self.regex_pattern_combo = QComboBox()
        self.regex_pattern_combo.addItems(
            [
                "None (Custom Regex)",
                "A-XXX-YYY-999",  
                "A-XXX-YYY-999_REV", 
                "A-XXX-YYY-999-REV", 
            ]
        )
        self.regex_pattern_combo.setCurrentIndex(1)
        self.regex_pattern_combo.currentIndexChanged.connect(
            self.update_regex_pattern
        )  # Connect to update function
        regex_select_layout.addWidget(self.regex_pattern_combo)
        regex_layout.addLayout(regex_select_layout)


        # --- Custom Regex Input ---
        custom_regex_layout = QHBoxLayout()
        custom_regex_layout.addWidget(QLabel("Custom Regex:"))
        self.custom_regex_input = QLineEdit()
        self.custom_regex_input.setPlaceholderText("Enter custom regex pattern")
        custom_regex_layout.addWidget(self.custom_regex_input)
        regex_layout.addLayout(custom_regex_layout)

        rename_layout.addLayout(regex_layout)

        # Connect the custom regex input to the update function
        self.custom_regex_input.textChanged.connect(self.update_masked_names)

        # -- Excel Options Sub-Group --
        self.excel_options_group = QGroupBox("Excel Options")
        excel_options_layout = QVBoxLayout()

        excel_file_layout = QHBoxLayout()
        excel_file_layout.addWidget(QLabel("Excel File:"))
        self.excel_path_input = QLineEdit()
        self.excel_path_input.setPlaceholderText("Select Excel file (.xlsx, .xlsm)")
        self.excel_path_button = QPushButton("Select...")
        self.excel_path_button.clicked.connect(self.select_excel_file)
        excel_file_layout.addWidget(self.excel_path_input)
        excel_file_layout.addWidget(self.excel_path_button)
        excel_options_layout.addLayout(excel_file_layout)

        excel_details_layout = QHBoxLayout()
        excel_details_layout.addWidget(QLabel("Tab Name:"))
        self.excel_tab_name_input = QLineEdit("Electrical ")  # Default
        excel_details_layout.addWidget(self.excel_tab_name_input)
        excel_details_layout.addSpacerItem(QSpacerItem(20, 10))
        excel_details_layout.addWidget(QLabel("Drawing No. Col:"))
        self.excel_drawing_num_input = QLineEdit("A")  # Default
        self.excel_drawing_num_input.setMaxLength(2)  # Allow AA etc.
        excel_details_layout.addWidget(self.excel_drawing_num_input)
        excel_details_layout.addSpacerItem(QSpacerItem(20, 10))
        excel_details_layout.addWidget(QLabel("Title Col:"))
        self.excel_drawing_title_input = QLineEdit("B")  # Default
        self.excel_drawing_title_input.setMaxLength(2)
        excel_details_layout.addWidget(self.excel_drawing_title_input)
        excel_options_layout.addLayout(excel_details_layout)

        excel_reload_layout = QHBoxLayout()
        self.excel_reload_button = QPushButton("Reload Excel Data and Apply to Table")
        self.excel_reload_button.clicked.connect(self.reload_and_apply_excel)
        excel_reload_layout.addStretch()
        excel_reload_layout.addWidget(self.excel_reload_button)
        excel_options_layout.addLayout(excel_reload_layout)

        self.excel_options_group.setLayout(excel_options_layout)
        rename_layout.addWidget(self.excel_options_group)

        # -- Rename Table --
        self.rename_table_view = PastingTableView()  # Use the enhanced view
        self.rename_model = QStandardItemModel(0, 2, self)
        self.rename_model.setHorizontalHeaderLabels(
            ["Masked Base Name", "Text to Append"]
        )  # Simplified header
        self.rename_table_view.setModel(self.rename_model)
        self.rename_table_view.verticalHeader().setVisible(False)
        self.rename_table_view.horizontalHeader().setStretchLastSection(True)
        self.rename_table_view.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )
        self.rename_table_view.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Interactive
        )
        self.rename_table_view.setMinimumHeight(150)
        rename_layout.addWidget(
            QLabel(
                "Edit 'Text to Append' below (if in Manual mode) or use Excel options."
            )
        )
        rename_layout.addWidget(self.rename_table_view)

        self.rename_group.setLayout(rename_layout)
        self.rename_group.setVisible(False)  # Initially hidden until PDF folder selected
        main_layout.addWidget(self.rename_group)

        # --- Metadata Editing Group ---
        # (This section remains unchanged)
        metadata_group = QGroupBox("Metadata Editing Options")
        metadata_main_layout = QVBoxLayout()
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setMinimumHeight(250)
        scroll_widget = QWidget()
        scroll_area.setWidget(scroll_widget)
        form_layout = QVBoxLayout(scroll_widget)
        self.field_widgets = {}
        self.field_widgets["title"] = self._create_metadata_field_widgets(
            form_layout,
            "Title",
            [
                "Leave Title Unchanged",
                "Use Specific Title",
                "Use Filename",
                "Use Filename (After First Space)",
                "Clear Title",
            ],
        )
        self.field_widgets["author"] = self._create_metadata_field_widgets(
            form_layout,
            "Author",
            ["Leave Unchanged", "Clear Author", "Set Specific Author"],
        )
        self.field_widgets["subject"] = self._create_metadata_field_widgets(
            form_layout,
            "Subject",
            ["Leave Unchanged", "Clear Subject", "Set Specific Subject"],
        )
        self.field_widgets["creator"] = self._create_metadata_field_widgets(
            form_layout,
            "Creator",
            ["Leave Unchanged", "Clear Creator", "Set Specific Creator"],
        )
        self.field_widgets["producer"] = self._create_metadata_field_widgets(
            form_layout,
            "Producer",
            ["Leave Unchanged", "Clear Producer", "Set Specific Producer"],
        )
        self.field_widgets["keywords"] = self._create_metadata_field_widgets(
            form_layout,
            "Keywords",
            ["Leave Unchanged", "Clear Keywords", "Set Specific Keywords"],
        

        )
        form_layout.addStretch()
        metadata_main_layout.addWidget(scroll_area)
        metadata_group.setLayout(metadata_main_layout)
        main_layout.addWidget(metadata_group)

        # --- Progress Bar and Log ---
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        main_layout.addWidget(self.progress_bar)
        self.log_area = QLineEdit("Status: Idle")
        self.log_area.setReadOnly(True)
        main_layout.addWidget(self.log_area)

        # --- Action Buttons ---
        button_layout = QHBoxLayout()
        self.start_button = QPushButton("Start Processing")
        self.start_button.clicked.connect(self.start_processing)
        self.stop_button = QPushButton("Stop")
        self.stop_button.clicked.connect(self.stop_processing)
        self.stop_button.setEnabled(False)
        button_layout.addWidget(self.start_button)
        button_layout.addWidget(self.stop_button)
        main_layout.addLayout(button_layout)

        # Set initial states
        self._update_all_input_states()
        self.toggle_output_folder()
        self._update_rename_mode()  # Set initial state of rename options

        # Initialize the custom regex field
        self.update_regex_pattern()

    # --- Helper for Creating Metadata UI Elements ---
    # (This method remains unchanged)
    def _create_metadata_field_widgets(self, parent_layout, field_name, combo_items):
        widgets = {}
        field_key = (
            field_name.lower()
        )  # Use lowercase for dict keys ('title', 'author')

        mode_layout = QHBoxLayout()
        widgets["mode_label"] = QLabel(f"{field_name} Action:")
        widgets["mode_combo"] = QComboBox()
        widgets["mode_combo"].addItems(combo_items)
        mode_layout.addWidget(widgets["mode_label"])
        mode_layout.addWidget(widgets["mode_combo"])
        mode_layout.addStretch()
        parent_layout.addLayout(mode_layout)

        specific_layout = QHBoxLayout()
        widgets["specific_label"] = QLabel(f"Specific {field_name}:")
        widgets["specific_edit"] = QLineEdit()
        specific_layout.addWidget(widgets["specific_label"])
        specific_layout.addWidget(widgets["specific_edit"])
        parent_layout.addLayout(specific_layout)

        # Connect signal to generic update function using the field_key
        widgets["mode_combo"].currentIndexChanged.connect(
            # Use lambda with default argument capture for key
            lambda idx, key=field_key: self._update_specific_input_state(key)
        )

        parent_layout.addSpacerItem(
            QSpacerItem(20, 10, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)
        )
        return widgets

    # --- UI State Update Functions ---
    # (These methods remain unchanged)
    def _update_specific_input_state(self, field_key):
        """Enable/disable the specific input based on combo box selection."""
        widgets = self.field_widgets[field_key]
        combo_text = widgets["mode_combo"].currentText()
        # Check if the combo text indicates setting a specific value
        is_specific_mode = "Specific" in combo_text
        widgets["specific_label"].setEnabled(is_specific_mode)
        widgets["specific_edit"].setEnabled(is_specific_mode)
        if not is_specific_mode:
            widgets["specific_edit"].clear()

    def _update_all_input_states(self):
        """Call update state for all fields."""
        for field_key in self.field_widgets.keys():
            self._update_specific_input_state(field_key)

    def toggle_output_folder(self):
        """Enable/disable output folder selection based on checkbox."""
        is_checked = self.edit_in_place_checkbox.isChecked()
        self.out_folder_input.setEnabled(not is_checked)
        self.out_folder_button.setEnabled(not is_checked)
        if is_checked:
            self.out_folder_input.clear()
            self.out_folder_input.setPlaceholderText(
                "Outputting to Input Folder"
            )  # Changed placeholder
        else:
            # Suggest default if input folder is set
            pdf_folder = self.pdf_folder_input.text()
            if pdf_folder and not self.out_folder_input.text():
                self.out_folder_input.setText(pdf_folder + "_processed")
            elif not pdf_folder:
                # Don't clear user input if they selected output first
                if not self.out_folder_input.text():
                    self.out_folder_input.setPlaceholderText("Select Input Folder first")
            else:
                # Keep existing text if user already selected one
                pass

    def _update_rename_mode(self):
        """Update UI elements based on the selected rename mode."""
        is_excel_mode = self.rename_mode_combo.currentText() == "Append from Excel"
        self.excel_options_group.setEnabled(is_excel_mode)

        # Update table column 1 editability
        self._set_table_append_column_editable(not is_excel_mode)

        if is_excel_mode:
            # Attempt to load and apply Excel data if switching to this mode
            # Only do this if files are already loaded in the table
            if self.rename_model.rowCount() > 0:
                self.reload_and_apply_excel()
        else:
            # Optional: Clear column 1 when switching to manual? Or leave existing data?
            # Let's leave existing data for now, user can delete/paste over.
            pass

    def _set_table_append_column_editable(self, editable: bool):
        """Sets the editable flag for all items in column 1."""
        self.rename_model.blockSignals(True)
        try:
            for row in range(self.rename_model.rowCount()):
                item = self.rename_model.item(row, 1)  # Column 1 (Append Text)
                if item:
                    flags = item.flags()
                    if editable:
                        flags |= Qt.ItemFlag.ItemIsEditable
                    else:
                        flags &= ~Qt.ItemFlag.ItemIsEditable
                    item.setFlags(flags)
        finally:
            self.rename_model.blockSignals(False)

    # --- File/Folder Selection Dialogs ---
    def select_pdf_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "Select PDF Input Folder")
        if folder_path:
            self.pdf_folder_input.setText(folder_path)
            # Update output folder suggestion/state *after* setting input
            self.toggle_output_folder()
            self.populate_rename_table(folder_path)  # Load files into table
            self.rename_group.setVisible(True)  # Show the rename group
            # Trigger rename mode update in case Excel mode is selected
            self._update_rename_mode()

    def select_output_folder(self):
        # Only allow selection if not outputting to input folder
        if not self.edit_in_place_checkbox.isChecked():
            folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
            if folder:
                self.out_folder_input.setText(folder)

    def select_excel_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx *.xlsm)"
        )
        if file_path:
            self.excel_path_input.setText(file_path)
            # Automatically try reloading when a new file is selected
            if self.rename_mode_combo.currentText() == "Append from Excel":
                self.reload_and_apply_excel()

    # --- Excel Data Handling ---
    def _load_excel_mapping(self):
        """Loads drawing number to title mapping from the specified Excel file."""
        excel_file = self.excel_path_input.text()
        tab_name = self.excel_tab_name_input.text()
        num_col_str = self.excel_drawing_num_input.text().strip().upper()
        title_col_str = self.excel_drawing_title_input.text().strip().upper()
        self.excel_mapping = {}  # Clear previous mapping

        if not excel_file or not os.path.exists(excel_file):
            self.log_area.setText("Status: Excel file not found or not selected.")
            return False
        if not tab_name:
            self.log_area.setText("Status: Excel Tab Name is required.")
            return False
        if not num_col_str or not title_col_str:
            self.log_area.setText("Status: Excel column letters are required.")
            return False

        try:
            # Convert column letters to indices (A=0, B=1, ..., Z=25, AA=26, ...)
            num_col_idx = openpyxl.utils.column_index_from_string(num_col_str) - 1
            title_col_idx = openpyxl.utils.column_index_from_string(title_col_str) - 1
        except ValueError:
            self.log_area.setText("Status: Invalid Excel column letters.")
            return False

        try:
            wb = openpyxl.load_workbook(
                excel_file, data_only=True, read_only=True
            )  # Read-only faster
            if tab_name not in wb.sheetnames:
                self.log_area.setText(f"Status: Excel worksheet '{tab_name}' not found.")
                return False
            ws = wb[tab_name]

            loaded_count = 0
            for row in ws.iter_rows(min_row=1, values_only=True):
                # Check if row has enough columns and cells are not None
                if (
                    len(row) > max(num_col_idx, title_col_idx)
                    and row[num_col_idx] is not None
                    and row[title_col_idx] is not None
                ):
                    drawing_number = str(row[num_col_idx]).strip()
                    drawing_title = str(row[title_col_idx]).strip()
                    if drawing_number:  # Need at least a drawing number
                        # Store the first title found for a given number
                        if drawing_number not in self.excel_mapping:
                            self.excel_mapping[drawing_number] = drawing_title
                            loaded_count += 1

            if not self.excel_mapping:
                self.log_area.setText(
                    "Status: No valid drawing number/title pairs found in Excel."
                )
                return False
            else:
                self.log_area.setText(f"Status: Loaded {loaded_count} mappings from Excel.")
                return True

        except Exception as e:
            self.log_area.setText(f"Status: Error reading Excel file: {e}")
            QMessageBox.critical(self, "Excel Error", f"Could not read Excel file:\n{e}")
            return False

    def _apply_excel_mapping_to_table(self):
        """Applies the loaded Excel mapping to the 'Text to Append' column."""
        if not self.excel_mapping:
            self.log_area.setText("Status: No Excel mapping loaded to apply.")
            # Ensure column is editable if mapping is empty/failed
            self._set_table_append_column_editable(True)
            return

        is_excel_mode = self.rename_mode_combo.currentText() == "Append from Excel"
        if not is_excel_mode:
            # Should not happen if called correctly, but safety check
            self._set_table_append_column_editable(True)
            return

        self.rename_model.blockSignals(True)
        found_count = 0
        try:
            for row in range(self.rename_model.rowCount()):
                item_masked = self.rename_model.item(row, 0)
                item_append = self.rename_model.item(row, 1)
                if not item_masked or not item_append:
                    continue

                original_filename = item_masked.data(Qt.UserRole)
                if not original_filename:
                    continue

                original_base, _ = os.path.splitext(original_filename)
                found_title = ""

                # Find the *first* matching drawing number from the mapping in the filename base
                # Iterate through mapping keys (drawing numbers)
                # This could be slow for huge mappings; consider optimizing if needed
                # (e.g., if drawing numbers have a fixed format, use regex)
                for drawing_number, drawing_title in self.excel_mapping.items():
                    # Simple substring check - adjust if more specific matching is needed
                    if drawing_number in original_base:
                        found_title = drawing_title
                        found_count += 1
                        break  # Use the first match

                item_append.setText(found_title)
                # Set flags (redundant if _update_rename_mode was called, but safe)
                flags = item_append.flags()
                flags &= ~Qt.ItemFlag.ItemIsEditable  # Ensure read-only in Excel mode
                item_append.setFlags(flags)

        finally:
            self.rename_model.blockSignals(False)
            # Emit dataChanged signal for the entire column to force view update
            if self.rename_model.rowCount() > 0:
                self.rename_model.dataChanged.emit(
                    self.rename_model.index(0, 1),
                    self.rename_model.index(self.rename_model.rowCount() - 1, 1),
                )
        self.log_area.setText(
            f"Status: Applied Excel mapping. Found titles for {found_count} files."
        )

    def reload_and_apply_excel(self):
        """Button action: Loads Excel data and applies it to the table."""
        if self.rename_mode_combo.currentText() != "Append from Excel":
            QMessageBox.information(
                self, "Mode Error", "Switch to 'Append from Excel' mode first."
            )
            return
        if self._load_excel_mapping():
            self._apply_excel_mapping_to_table()
        else:
            # Load failed, ensure column 1 is cleared and editable status is correct
            self._set_table_append_column_editable(
                False
            )  # Still Excel mode, so not editable
            # Clear column 1 if load failed?
            self.rename_model.blockSignals(True)
            for row in range(self.rename_model.rowCount()):
                item = self.rename_model.item(row, 1)
                if item:
                    item.setText("")
            self.rename_model.blockSignals(False)
            if self.rename_model.rowCount() > 0:
                self.rename_model.dataChanged.emit(
                    self.rename_model.index(0, 1),
                    self.rename_model.index(self.rename_model.rowCount() - 1, 1),
                )

    # --- Populate Rename Table (Handles Masking and potentially Excel) ---
    def populate_rename_table(self, folder_path):
        self.rename_model.removeRows(0, self.rename_model.rowCount())  # Clear existing rows
        self.excel_mapping = {}  # Clear old mapping on folder change
        try:
            pdf_files = sorted(
                [
                    f
                    for f in os.listdir(folder_path)
                    if f.lower().endswith(".pdf")
                    and os.path.isfile(os.path.join(folder_path, f))
                ]
            )

            if not pdf_files:
                self.log_area.setText("Status: No PDF files found in selected folder.")
                self.rename_group.setVisible(False)  # Hide if no files
                return

            # Get current regex pattern from input field
            regex_pattern = self.custom_regex_input.text()

            self.rename_model.setRowCount(len(pdf_files))
            for row, filename in enumerate(pdf_files):
                original_base, _ = os.path.splitext(filename)

                # Calculate masked name using the regex logic
                masked_name = self._calculate_masked_name(original_base, regex_pattern)

                # Column 0: Masked Base Name (DisplayRole), Original Filename (UserRole)
                item_masked_name = QStandardItem(masked_name)
                item_masked_name.setData(filename, Qt.UserRole)  # Store original filename
                item_masked_name.setFlags(
                    item_masked_name.flags() & ~Qt.ItemFlag.ItemIsEditable
                )  # Make read-only
                self.rename_model.setItem(row, 0, item_masked_name)

                # Column 1: Text to Append (Editable or populated by Excel later)
                item_append = QStandardItem("")  # Initially empty
                self.rename_model.setItem(row, 1, item_append)

            self.rename_table_view.resizeColumnToContents(1)
            self.log_area.setText(f"Status: Loaded {len(pdf_files)} PDF files. Ready.")

            # Now that table is populated, check current mode and apply Excel if needed
            self._update_rename_mode()  # This will trigger reload/apply if in Excel mode

        except Exception as e:
            QMessageBox.critical(
                self, "Error Listing Files", f"Could not read PDF folder contents:\n{e}"
            )
            self.rename_group.setVisible(False)

    # --- Helper to Calculate Masked Name using Regex ---
    def _calculate_masked_name(self, original_base, regex_pattern):
        """Calculates the masked name based on the provided regex."""
        if not original_base:
            return ""

        if not regex_pattern or regex_pattern.strip() == "":
            return original_base  # Return original if no regex

        try:
            match = re.search(regex_pattern, original_base)
            if match:
                return match.group(0)  # Return the entire matched string
            else:
                return ""  # Regex not found, mask fails
        except re.error as e:
            print(f"Regex error: {e}")
            return ""  # Regex error, mask fails

    # --- Slot to Update Masked Names in Table ---
    def update_masked_names(self):
        """Recalculates and updates the first column when regex changes."""
        if not self.rename_group.isVisible() or self.rename_model.rowCount() == 0:
            return  # Nothing to update

        # Get current regex pattern from input field
        regex_pattern = self.custom_regex_input.text()

        self.rename_model.blockSignals(True)  # Block signals while updating
        try:
            for row in range(self.rename_model.rowCount()):
                item = self.rename_model.item(row, 0)
                if item:
                    original_filename = item.data(Qt.UserRole)
                    if original_filename:
                        original_base, _ = os.path.splitext(original_filename)
                        new_masked_name = self._calculate_masked_name(
                            original_base, regex_pattern
                        )
                        item.setData(new_masked_name, Qt.DisplayRole)  # Update display text
        finally:
            self.rename_model.blockSignals(False)  # Unblock signals
            self.rename_model.dataChanged.emit(
                self.rename_model.index(0, 0),
                self.rename_model.index(self.rename_model.rowCount() - 1, 0),
            )  # Emit dataChanged signal to force view update


    # --- Slot to Update Regex Pattern from ComboBox ---
    def update_regex_pattern(self):
        """Updates the custom regex input based on the combo box selection."""
        selected_pattern = self.regex_pattern_combo.currentText()
        if selected_pattern == "A-XXX-YYY-999":
            self.custom_regex_input.setText(r"[A-Za-z]{1}-[A-Za-z]{3}-[A-Za-z]{3}-\d{3}")  # Added regex
        elif selected_pattern == "A-XXX-YYY-999_REV":
            self.custom_regex_input.setText(r"\b[A-Z]-[A-Z]{3}-[A-Z]{3}-\d{3}_[A-Z0-9]{1,2}\b")
        elif selected_pattern == "A-XXX-YYY-999-REV":
            self.custom_regex_input.setText(r"\b[A-Z]-[A-Z]{3}-[A-Z]{3}-\d{3}-[A-Z0-9]{1,2}\b")
        else:  # "None (Custom Regex)"
            self.custom_regex_input.clear()

        # Trigger the name update to apply the selected regex
        self.update_masked_names()

    # --- Worker Interaction ---
    # (update_log, update_progress remain unchanged)
    def update_log(self, message):
        self.log_area.setText(message)

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    # --- Slot for Worker Finished Signal ---
    # (This method remains unchanged)
    def processing_finished(self, planned_renames, edited_succeeded, edited_failed):
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        QMessageBox.information(
            self,
            "Processing Complete",
            f"Finished.\n\n"
            f"Output Renames Planned: {planned_renames}\n"
            f"Files Processed (Metadata): {edited_succeeded} succeeded, {edited_failed} failed.",
        )
        final_status = (
            f"Status: Completed (Renames Planned: {planned_renames} | "
            f"Processed: {edited_succeeded}✓ {edited_failed}✗)"
        )
        self.log_area.setText(final_status)
        self.worker = None

    # --- Start Processing (Extracts Data from Table) ---
    def start_processing(self):
        pdf_folder = self.pdf_folder_input.text()

        if self.edit_in_place_checkbox.isChecked():
            output_folder = pdf_folder
        else:
            output_folder = self.out_folder_input.text()

        # --- Input Validation ---
        if not pdf_folder or not os.path.isdir(pdf_folder):
            QMessageBox.warning(self, "Input Error", "Select a valid PDF Input folder.")
            return
        if not self.edit_in_place_checkbox.isChecked() and not output_folder:
            QMessageBox.warning(
                self, "Input Error", "Select an Output folder or check 'Output to Input Folder'."
            )
            return
        # Allow output == input only if checkbox is checked
        if (
            not self.edit_in_place_checkbox.isChecked()
            and os.path.abspath(pdf_folder) == os.path.abspath(output_folder)
        ):
            QMessageBox.warning(
                self,
                "Input Error",
                "Output folder cannot be the same as Input folder unless 'Output to Input Folder' is checked.",
            )
            return

        # --- Validate Excel Settings if in Excel Mode ---
        if (
            self.rename_group.isVisible()
            and self.rename_mode_combo.currentText() == "Append from Excel"
        ):
            excel_file = self.excel_path_input.text()
            tab_name = self.excel_tab_name_input.text()
            num_col_str = self.excel_drawing_num_input.text().strip().upper()
            title_col_str = self.excel_drawing_title_input.text().strip().upper()
            if not excel_file or not os.path.exists(excel_file):
                QMessageBox.warning(
                    self,
                    "Input Error",
                    "Excel file not found or not selected for 'Append from Excel' mode.",
                )
                return
            if not tab_name:
                QMessageBox.warning(
                    self,
                    "Input Error",
                    "Excel Tab Name cannot be empty for 'Append from Excel' mode.",
                )
                return
            try:
                openpyxl.utils.column_index_from_string(num_col_str)
                openpyxl.utils.column_index_from_string(title_col_str)
            except ValueError:
                QMessageBox.warning(
                    self,
                    "Input Error",
                    "Invalid Excel column letters provided for 'Append from Excel' mode.",
                )
                return
            # Check if mapping was actually loaded successfully earlier
            # Check if mapping was actually loaded successfully earlier
            if not self.excel_mapping:
                reply = QMessageBox.question(
                    self,
                    "Confirm Action",
                    "Excel mapping is currently empty or failed to load previously.\n"
                    "No text will be appended from Excel.\n\n"
                    "Do you want to continue?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,  # Default to No
                )
                if reply == QMessageBox.StandardButton.No:
                    return  # Stop processing if user cancels
                # If user says Yes, processing continues, but excel_mapping is empty,
                # so no text will be appended from it.

        # --- Extract Renaming Data from Table ---
        rename_data = []
        all_append_texts_empty = True  # Track if any append text exists
        if self.rename_group.isVisible() and self.rename_model.rowCount() > 0:
            for row in range(self.rename_model.rowCount()):
                item_masked_name = self.rename_model.item(row, 0)  # Item for first column
                item_append = self.rename_model.item(row, 1)  # Item for second column

                if item_masked_name and item_append:
                    original_filename = item_masked_name.data(
                        Qt.UserRole
                    )  # Get original from UserRole
                    masked_base_name = item_masked_name.text()  # Get current masked name from DisplayRole
                    text_to_append = item_append.text().strip()  # Get text from table (manual or Excel populated)

                    if original_filename:  # Ensure we have the original name
                        rename_data.append(
                            (original_filename, masked_base_name, text_to_append)
                        )
                        if text_to_append:
                            all_append_texts_empty = True
        
        
                            all_append_texts_empty = (
                                False  # Found some text to append
                            )
                    else:
                        print(f"Warning: Missing original filename for row {row}")  # Should not happen
        elif self.rename_group.isVisible() and self.rename_model.rowCount() == 0:
            # If rename group is visible but no files loaded
            QMessageBox.warning(
                self, "Input Error", "No PDF files loaded in the table to process."
            )
            return

        # --- Confirmation Dialogs ---
        # Only ask confirmation if in Manual mode and all append texts are empty
        if (
            self.rename_group.isVisible()
            and self.rename_mode_combo.currentText() == "Manual / Paste Table"
            and all_append_texts_empty
            and self.rename_model.rowCount() > 0
        ):
            reply = QMessageBox.question(
                self,
                "Confirm Action",
                "No text was entered or pasted into the 'Text to Append' column.\n"
                "The output filenames will only be based on the 'Masked Base Name'.\n\n"
                "Do you want to continue?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes,
            )  # Default to Yes
            if reply == QMessageBox.StandardButton.No:
                return

        # --- Metadata settings ---
        # (This section remains unchanged)
        metadata_settings = {}
        for field_key, widgets in self.field_widgets.items():
            mode_text = widgets["mode_combo"].currentText()
            specific_value = widgets["specific_edit"].text()
            mode_internal = "Leave"
            if "Specific" in mode_text:
                mode_internal = "Specific"
            elif "Clear" in mode_text:
                mode_internal = "Clear"
            elif "Filename (After First Space)" in mode_text:
                mode_internal = "FilenameAfterSpace"
            elif "Filename" in mode_text:
                mode_internal = "Filename"
            metadata_settings[field_key] = {
                "mode": mode_internal,
                "value": specific_value if mode_internal == "Specific" else None,
            }

        # --- Confirmation for Outputting to Input Folder ---
        # (This section remains unchanged)
        if self.edit_in_place_checkbox.isChecked():
            reply = QMessageBox.question(
                self,
                "Confirmation",
                "You have chosen 'Output to Input Folder'.\n"
                "Processed files will be saved in the same folder as the input PDFs, potentially overwriting existing files if names match.\n\n"
                "Do you want to continue?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )  # Default to No
            if reply == QMessageBox.StandardButton.No:
                return

        # --- Start Worker (Pass updated rename_data structure) ---
        # The worker doesn't need to know about Excel, it just gets the final data
        self.worker = CombinedWorker(
            pdf_folder=pdf_folder,
            output_folder=output_folder,
            rename_data=rename_data,  # Contains original, masked_base, and append_text
            metadata_settings=metadata_settings,
        )
        self.worker.progress.connect(self.update_progress)
        self.worker.log_message.connect(self.update_log)
        self.worker.finished.connect(self.processing_finished)

        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.progress_bar.setValue(0)
        self.log_area.setText("Status: Starting...")
        self.worker.start()

    # --- Stop Processing and Close Event ---
    # (These methods remain unchanged)
    def stop_processing(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log_area.setText("Status: Stopping...")
            # Keep stop button enabled until worker confirms finished

    def closeEvent(self, event):
        if self.worker and self.worker.isRunning():
            reply = QMessageBox.question(
                self,
                "Confirm Exit",
                "Processing is ongoing. Stop and exit?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.stop_processing()
                # Don't wait here, let the app close. Worker should handle stop signal.
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()


# --- Run Application ---
if __name__ == "__main__":
    # Ensure required libraries are installed
    try:
        import openpyxl  # Added back
        import pypdf
        from PySide6.QtWidgets import QApplication
    except ImportError as e:
        print(f"Error: Missing required library. Please install dependencies.")
        print(f"Missing: {e.name}")
        # Updated install command
        print("Try running: pip install pypdf PySide6 openpyxl")
        sys.exit(1)

    app = QApplication(sys.argv)
    window = CombinedPdfProcessorApp()
    window.show()
    sys.exit(app.exec())
